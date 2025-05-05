import openpyxl
from tkinter import filedialog, Tk

def exchange_columns(file_path):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Intercambiar los nombres de las columnas
    sheet['A1'].value, sheet['A1'].value = sheet['B1'].value, sheet['A1'].value

    # Intercambiar los valores entre las columnas
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
        clave_catastral = row[0].value
        clave_municipio = row[1].value
        row[0].value = clave_municipio
        row[1].value = clave_catastral

    # Guardar los cambios en un nuevo archivo Excel
    save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    wb.save(save_path)

    print("¡Intercambio completado y archivo guardado con éxito!")

def main():
    # Abrir el cuadro de diálogo para seleccionar el archivo Excel
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    root.destroy()  # Cerrar la ventana principal

    if file_path:
        exchange_columns(file_path)
    else:
        print("No se seleccionó ningún archivo.")

if __name__ == "__main__":
    main()
