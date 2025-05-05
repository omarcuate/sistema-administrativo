import openpyxl
from tkinter import filedialog, Tk

def process_excel(file_path):
    # Cargar el archivo Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Crear dos nuevos libros de trabajo
    wb_correct = openpyxl.Workbook()
    wb_incorrect = openpyxl.Workbook()

    # Crear dos hojas de cálculo para cada libro de trabajo
    sheet_correct = wb_correct.active
    sheet_incorrect = wb_incorrect.active

    # Copiar el encabezado en ambas hojas de cálculo
    for cell in sheet[1]:
        sheet_correct.cell(row=1, column=cell.column, value=cell.value)
        sheet_incorrect.cell(row=1, column=cell.column, value=cell.value)

    # Separar las claves con exactamente 16 caracteres y las demás
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        clave_catastral = str(row[0].value)
        new_row = [cell.value for cell in row]
        if len(clave_catastral) == 16:
            sheet_correct.append(new_row)
        else:
            sheet_incorrect.append(new_row)

    # Abrir el cuadro de diálogo para seleccionar la ruta de almacenamiento para los archivos generados
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal
    save_path_correct = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Guardar archivo de claves correctas como")
    save_path_incorrect = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Guardar archivo de claves incorrectas como")
    root.destroy()  # Cerrar la ventana principal

    # Guardar los archivos
    if save_path_correct:
        wb_correct.save(save_path_correct)
        print("Archivo de claves municipio correctas guardado con éxito:", save_path_correct)
    else:
        print("No se seleccionó una ubicación para guardar el archivo de claves municipio correctas.")

    if save_path_incorrect:
        wb_incorrect.save(save_path_incorrect)
        print("Archivo de claves municipio incorrectas guardado con éxito:", save_path_incorrect)
    else:
        print("No se seleccionó una ubicación para guardar el archivo de claves municipio incorrectas.")

    # Mostrar mensaje de éxito
    print("Proceso completado. Archivos procesados y guardados correctamente.")

def main():
    # Abrir el cuadro de diálogo para seleccionar el archivo Excel
    root = Tk()
    root.withdraw()  # Ocultar la ventana principal
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], title="Seleccionar archivo Excel")
    root.destroy()  # Cerrar la ventana principal

    if file_path:
        process_excel(file_path)
    else:
        print("No se seleccionó ningún archivo.")

if __name__ == "__main__":
    main()
